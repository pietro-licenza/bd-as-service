"""
API routes for image processing endpoints.
"""
import os
import json
import tempfile
import logging
import re
from typing import List
from datetime import datetime
from fastapi import UploadFile, File, Depends, APIRouter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session # Adicionado para Supabase

from app.models.schemas import BatchResponse, BatchProductResponse
from app.models.entities import ScrapingLog # Adicionado para Supabase
from app.shared.gemini_client import send_to_gemini, generate_product_images_with_gemini
from app.core.auth import get_current_user
from app.core.database import get_db # Adicionado para Supabase

# Configure logging
logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# Initialize router
router = APIRouter()

# Ensure exports directory exists
os.makedirs("exports", exist_ok=True)


def generate_excel_report(products_data: List[dict], filename: str) -> str:
    """
    Generate Excel report from products data.
    
    Args:
        products_data: List of product dictionaries
        filename: Output filename
        
    Returns:
        Path to generated Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos Processados"
    
    # Define headers
    headers = ["ID", "Nome do Produto", "Preço", "EAN", "Especificações", "Descrição", "Imagens Geradas", "Qtd Imagens"]
    
    # Style headers
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for row_num, product in enumerate(products_data, 2):
        ws.cell(row=row_num, column=1, value=product.get("product_id", ""))
        ws.cell(row=row_num, column=2, value=product.get("nome_produto", ""))
        ws.cell(row=row_num, column=3, value=product.get("preco", ""))
        ws.cell(row=row_num, column=4, value=product.get("ean", ""))
        
        # Especificações (join list)
        specs = product.get("especificacoes", [])
        ws.cell(row=row_num, column=5, value="\n".join(specs) if specs else "")
        
        ws.cell(row=row_num, column=6, value=product.get("descricao", ""))
        
        # URLs (join list with comma)
        urls = product.get("generated_images_urls", [])
        ws.cell(row=row_num, column=7, value=", ".join(urls) if urls else "")
        
        ws.cell(row=row_num, column=8, value=product.get("num_generated", 0))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 80
    ws.column_dimensions['H'].width = 12
    
    # Set row heights and wrap text
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Save file
    filepath = os.path.join("exports", filename)
    wb.save(filepath)
    logger.info(f"✅ Excel gerado: {filepath}")
    
    return filepath


@router.post("/process-batch")
async def process_batch_endpoint(
    files: List[UploadFile] = File(...), 
    user = Depends(get_current_user),
    db: Session = Depends(get_db) # Injeção da conexão Supabase
) -> BatchResponse:
    """
    Process multiple products in batch and log usage to Supabase.
    """
    logger.info(f"🚀 Iniciando processamento em lote de {len(files)} imagens para o usuário {user.username}")
    
    # Group files by product based on filename pattern
    product_groups = {}
    unmatched_files = []
    
    for file in files:
        filename = file.filename
        product_id = None
        
        match = re.match(r'(product[_\s]?\d+)', filename.lower())
        if match:
            product_id = match.group(1).replace(' ', '_')
            if product_id not in product_groups:
                product_groups[product_id] = []
            product_groups[product_id].append(file)
        else:
            unmatched_files.append(file)
    
    if unmatched_files:
        logger.warning(f"⚠️  {len(unmatched_files)} arquivo(s) sem padrão 'productX_'. Agrupando a cada 3 imagens.")
        images_per_product = 3
        for i in range(0, len(unmatched_files), images_per_product):
            product_key = f"product_{len(product_groups) + 1}"
            product_groups[product_key] = unmatched_files[i:i + images_per_product]
    
    logger.info(f"📦 {len(product_groups)} produto(s) identificado(s)")
    
    all_temp_files = []
    results = []
    total_images = 0
    product_counter = 1
    
    try:
        for product_key, product_files in product_groups.items():
            logger.info(f"\n{'='*60}")
            logger.info(f"🔄 Processando Produto #{product_counter} ({product_key})")
            logger.info(f"{'='*60}")
            
            temp_files = []
            image_paths = []
            filenames = []
            
            try:
                logger.info(f"📸 Salvando {len(product_files)} imagem(ns) temporariamente...")
                for file in product_files:
                    with tempfile.NamedTemporaryFile(
                        delete=False, 
                        suffix=os.path.splitext(file.filename)[1], 
                        mode="wb"
                    ) as temp_file:
                        content = await file.read()
                        temp_file.write(content)
                        temp_file.flush()
                        temp_file_path = temp_file.name
                        temp_files.append(temp_file_path)
                        all_temp_files.append(temp_file_path)

                    image_paths.append(temp_file_path)
                    filenames.append(file.filename)

                logger.info(f"🤖 Enviando imagens para Gemini (extração de dados)...")
                gemini_result = send_to_gemini(image_paths, [], [], [])
                gemini_response_text = gemini_result["gemini_response"]
                product_image_path = gemini_result["product_image_path"]
                
                try:
                    gemini_data = json.loads(gemini_response_text.replace("```json", "").replace("```", "").strip())
                    product_name = gemini_data.get("nome_produto", "produto")
                except:
                    product_name = "produto"
                    gemini_data = {}
                
                logger.info(f"🎨 Gerando imagens com fundo branco (Gemini 2.0)...")
                generated_result = generate_product_images_with_gemini(product_image_path, product_name, product_counter)
                
                response_data = {
                    "nome_produto": gemini_data.get("nome_produto"),
                    "preco": gemini_data.get("preco"),
                    "ean": gemini_data.get("ean"),
                    "especificacoes": gemini_data.get("especificacoes", []),
                    "descricao": gemini_data.get("descricao"),
                    "generated_images_urls": generated_result["public_urls"],
                    "num_generated": generated_result["num_generated"]
                }
                
                results.append(BatchProductResponse(
                    product_id=product_counter,
                    num_images=len(product_files),
                    filenames=filenames,
                    gemini_response=json.dumps(response_data, ensure_ascii=False, indent=2)
                ))

                # --- REGISTRO NO SUPABASE ---
                try:
                    log_entry = ScrapingLog(
                        user_id=user.id, # Vincula ao Victor logado
                        loja="processamento_lote",
                        url_count=len(product_files),
                        total_tokens=2500,  # Valor médio de tokens/produto
                        total_cost_brl=0.20 # Custo estimado por produto processado
                    )
                    db.add(log_entry)
                    db.commit()
                    logger.info(f"📊 Log de scraping salvo no banco para {user.username}")
                except Exception as db_err:
                    db.rollback()
                    logger.error(f"❌ Erro ao salvar log no Supabase: {db_err}")
                
                total_images += len(product_files)
                product_counter += 1

            except Exception as e:
                logger.error(f"❌ Erro ao processar Produto #{product_counter}: {str(e)}")
                results.append(BatchProductResponse(
                    product_id=product_counter,
                    num_images=len(product_files),
                    filenames=[f.filename for f in product_files],
                    gemini_response="",
                    error=str(e)
                ))
                product_counter += 1

        # Finalização e Geração de Excel
        logger.info(f"🎉 Processamento concluído! {len(results)} produto(s) processado(s)")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"produtos_{timestamp}.xlsx"
        
        excel_data = []
        for result in results:
            if not result.error:
                try:
                    data = json.loads(result.gemini_response)
                    data["product_id"] = result.product_id
                    excel_data.append(data)
                except: pass
        
        excel_path = generate_excel_report(excel_data, excel_filename) if excel_data else None
        excel_url = f"/exports/{excel_filename}" if excel_path else None
        
        response = BatchResponse(
            total_products=len(product_groups),
            total_images=total_images,
            products=results
        )
        
        response_dict = response.dict()
        response_dict["excel_download_url"] = excel_url
        
        return response_dict

    finally:
        # Limpeza de arquivos temporários
        for temp_file_path in all_temp_files:
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)