"""
Gerador de Excel específico para o Kit Builder.
Igual ao generate_standard_excel do shared, mas com:
- Suporte nativo a múltiplos kits (uma linha por kit)
- Coluna extra ao final: URLS ORIGINAIS (URLs dos produtos do kit, separadas por vírgula)
"""
import os
import logging
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

KIT_BRAND_COLOR = "7C3AED"


def _parse_price(price_str: Any) -> float:
    """Converte string de preço (R$ 1.234,56) para float (1234.56)."""
    if isinstance(price_str, (int, float)):
        return float(price_str)
    if not price_str or not isinstance(price_str, str):
        return 0.0
    try:
        clean = price_str.replace("R$", "").replace(".", "").replace(",", ".").strip()
        return float(clean)
    except Exception:
        return 0.0


def generate_kit_excel(kits: List[Dict], filename: str, output_dir: str) -> str:
    """
    Gera um Excel com todos os kits fornecidos, uma linha por kit.

    Colunas padrão (A–W) idênticas ao generate_standard_excel:
      A  NOME DO PRODUTO
      B  TÍTULO DO PRODUTO
      C  EAN
      D  URL IMAGENS
      E  LARGURA (CM)
      F  COMPRIMENTO (CM)
      G  ALTURA (CM)
      H  DIMENSÕES (LXCXA)
      I  PESO (KG)
      J  PREÇO LOJA
      K  DESCONTO
      L  FRETE
      M  TARIFA
      N  PREÇO LOJA CUSTO
      O  PREÇO ANÚNCIO
      P  TESTE ARREDONDAMENTO
      Q  LUCRO %
      R  LUCRO % ARREDONDAMENTO
      S  LUCRO
      T  LUCRO ARREDONDAMENTO
      U  DESCRIÇÃO
      V  MARCA
      W  MODELO
      X  URLS ORIGINAIS   ← exclusiva do kit builder

    Coluna Y é reservada para "% LUCRO DESEJADO" (igual ao padrão).
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    filepath = os.path.join(output_dir, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    headers = [
        "NOME DO PRODUTO",        # A  1
        "TÍTULO DO PRODUTO",      # B  2
        "EAN",                    # C  3
        "URL IMAGENS",            # D  4
        "LARGURA (CM)",           # E  5
        "COMPRIMENTO (CM)",       # F  6
        "ALTURA (CM)",            # G  7
        "DIMENSÕES (LXCXA)",      # H  8
        "PESO (KG)",              # I  9
        "PREÇO LOJA",             # J  10
        "DESCONTO",               # K  11
        "FRETE",                  # L  12
        "TARIFA",                 # M  13
        "PREÇO LOJA CUSTO",       # N  14
        "PREÇO ANÚNCIO",          # O  15
        "TESTE ARREDONDAMENTO",   # P  16
        "LUCRO %",                # Q  17
        "LUCRO % ARREDONDAMENTO", # R  18
        "LUCRO",                  # S  19
        "LUCRO ARREDONDAMENTO",   # T  20
        "DESCRIÇÃO",              # U  21
        "MARCA",                  # V  22
        "MODELO",                 # W  23
        "URLS ORIGINAIS",         # X  24
    ]

    header_fill = PatternFill(start_color=KIT_BRAND_COLOR, end_color=KIT_BRAND_COLOR, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    standard_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Coluna Z: % Lucro Desejado (coluna Y — índice 25)
    profit_target_col = 25  # Y
    ws.cell(row=1, column=profit_target_col, value="% LUCRO DESEJADO").fill = header_fill
    ws.cell(row=1, column=profit_target_col).font = header_font
    ws.cell(row=2, column=profit_target_col, value=0.15).number_format = "0%"

    for idx, kit in enumerate(kits, 2):
        price_val = _parse_price(kit.get("preco", "0"))
        imgs = kit.get("image_urls", [])
        img_str = ", ".join(imgs) if isinstance(imgs, list) else str(imgs)
        urls_originais = kit.get("url_original", "")

        ws.row_dimensions[idx].height = 15

        # Dados básicos
        ws.cell(row=idx, column=1,  value=kit.get("titulo", "")).alignment = standard_align
        ws.cell(row=idx, column=2,  value="").alignment = standard_align
        ws.cell(row=idx, column=3,  value=kit.get("ean", "")).alignment = standard_align
        ws.cell(row=idx, column=4,  value=img_str).alignment = standard_align

        # Dimensões
        ws.cell(row=idx, column=5,  value=kit.get("largura_cm", "")).alignment = center_align
        ws.cell(row=idx, column=6,  value=kit.get("comprimento_cm", "")).alignment = center_align
        ws.cell(row=idx, column=7,  value=kit.get("altura_cm", "")).alignment = center_align
        ws.cell(row=idx, column=8,  value=kit.get("dimensoes_lca", "")).alignment = center_align
        ws.cell(row=idx, column=9,  value=kit.get("peso_kg", "")).alignment = center_align

        # Financeiro
        c_price = ws.cell(row=idx, column=10, value=price_val)
        c_price.number_format = "#,##0.00"
        c_price.alignment = standard_align

        ws.cell(row=idx, column=11, value=0).number_format    = "#,##0.00"
        ws.cell(row=idx, column=12, value=50).number_format   = "#,##0.00"
        ws.cell(row=idx, column=13, value=0.115).number_format = "0.00%"

        ws.cell(row=idx, column=14, value=f"=J{idx}-K{idx}").number_format                          = "#,##0.00"
        ws.cell(row=idx, column=15, value=f"=(N{idx}+L{idx})/(1-M{idx}-$Y$2)").number_format        = "#,##0.00"
        ws.cell(row=idx, column=16, value=f"=CEILING(O{idx}, 10)-0.1").number_format                = "#,##0.00"
        ws.cell(row=idx, column=17, value=f"=(O{idx}*(1-M{idx})-L{idx}-N{idx})/O{idx}").number_format = "0.00%"
        ws.cell(row=idx, column=18, value=f"=(P{idx}*(1-M{idx})-L{idx}-N{idx})/P{idx}").number_format = "0.00%"
        ws.cell(row=idx, column=19, value=f"=O{idx}*(1-M{idx})-L{idx}-N{idx}").number_format        = "#,##0.00"
        ws.cell(row=idx, column=20, value=f"=P{idx}*(1-M{idx})-L{idx}-N{idx}").number_format        = "#,##0.00"

        # Texto
        ws.cell(row=idx, column=21, value=kit.get("descricao", "")).alignment = standard_align
        ws.cell(row=idx, column=22, value=kit.get("marca", "Brazil Home Living")).alignment = standard_align
        ws.cell(row=idx, column=23, value=kit.get("modelo", "Kit")).alignment = standard_align

        # URLS ORIGINAIS — exclusivo kit builder
        ws.cell(row=idx, column=24, value=urls_originais).alignment = standard_align

    # Larguras das colunas
    widths = {
        1: 45, 2: 45, 3: 20, 4: 25, 5: 15, 6: 15, 7: 15, 8: 20, 9: 15,
        10: 15, 11: 12, 12: 12, 13: 12, 14: 20, 15: 20, 16: 22, 17: 15,
        18: 25, 19: 12, 20: 22, 21: 60, 22: 20, 23: 15, 24: 50, 25: 25,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(filepath)
    logger.info(f"✅ Kit Excel gerado: {filepath} ({len(kits)} kit(s))")
    return filepath
